from openpyxl import Workbook
from pathlib import Path
import os
from os import path
from utils import add_configuration, convert_url, create_folder, reading_json, create_json

def create_arq(instance, datas):
    ws = instance.create_sheet("Caixas",0)
    ws2 = instance.create_sheet("Pastas",-1)
    

    # Plan 01
    ws['A1'] = "Codigo"
    ws['B1'] = "Galpão"
    ws['C1'] = "Prateleira"
    ws['D1'] = "Cliente"

    ws2['A1'] = "Codigo"
    ws2['B1'] = "Departamento"
    ws2['C1'] = "Assunto"
    ws2['D1'] = "Indexação"
    ws2['E1'] = "Data_Inicio"
    ws2['F1'] = "Data_Final"
    ws2['G1'] = "Obs"
    ws2['H1'] = "Caixa"

    row = 2
    row2 = 2
    for data in datas:
        ws[f'A{row}'] = data['index']
        ws[f'B{row}'] = data['galpao']
        ws[f'C{row}'] = data['prateleira']
        ws[f'D{row}'] = data['client']
        row += 1
        for box in data['boxs']:
            ws2[f'A{row2}'] = box['index']
            ws2[f'B{row2}'] = box['department']
            ws2[f'C{row2}'] = box['subject']
            ws2[f'D{row2}'] = box['indexing']
            ws2[f'E{row2}'] = box['date_start']
            ws2[f'F{row2}'] = box['date_end']
            ws2[f'G{row2}'] = box['obs']
            ws2[f'H{row2}'] = data['index']
           
            row2 += 1


    # # Save the file
    instance.save(convert_url("./scan.xlsx"))
    add_configuration()


def add_xlsx():
    wb = Workbook()

    fileObj = Path(convert_url("./config/boxs.json"))

    if fileObj.is_file():
        datas = reading_json(convert_url('./config/boxs'))
        create_arq(wb, datas)

