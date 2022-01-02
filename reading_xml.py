from openpyxl import load_workbook
from utils import convert_url, trim

def get_end_values(instance, sheet, index):
    ws = instance[sheet]

    cells = 2
    location = ""
    while True:

        if ws[f'A{cells}'].value:
            if trim( ws[f'A{cells}'].value) == index:
                
                location = cells-1

            cells = cells + 1
        else:
            break
    
    return index,location, (cells - 2)



def update_xlsx(ult_box):
    wb =  load_workbook(convert_url("./scan.xlsx"))

    ws = wb['Caixas']
    ws2 = wb['Pastas']
    cells = 2

    datas = list()

    try:
        indx,loc, cel = get_end_values(wb, 'Caixas', ult_box['index'])
        indx2,loc2, cel2 = get_end_values(wb, 'Pastas', ult_box['boxs']['index'])
    except:
        indx = ""
        loc = 0
        cel = 0
        indx2 = ""
        loc2 = 0
        cel2 = 0

    if trim(ws2[f'H{(loc2+2)}'].value) == indx:
        cells = loc + 1
    else:
        cells = loc + 2

    while True:

        if ws[f'A{cells}'].value:

            datas.append({
                "index": trim(ws[f'A{cells}'].value),
                "galpao": trim(ws[f'B{cells}'].value),
                "prateleira": trim(ws[f'C{cells}'].value),
                "client": trim(ws[f'D{cells}'].value),
                "boxs": []
            })

            cells = cells + 1
        else:
            break
    
   
    
    cells2 = loc2 + 2

    fold = dict()
    while True:

        if ws2[f'A{cells2}'].value:
            try:
                val = fold[trim(ws2[f'H{cells2}'].value)]
            except:
                val = list()

            val.append({
                "index": trim(ws2[f'A{cells2}'].value),
                "department": trim(ws2[f'B{cells2}'].value),
                "subject": trim(ws2[f'C{cells2}'].value),
                "indexing": trim(ws2[f'D{cells2}'].value),
                "date_start": trim(ws2[f'E{cells2}'].value),
                "date_end": trim(ws2[f'F{cells2}'].value),
                "obs": trim(ws2[f'G{cells2}'].value),
            })

            fold[trim(ws2[f'H{cells2}'].value)] = val

            cells2 = cells2 + 1
        else:
            break

    for dt in datas:
        dt['boxs'] = fold[dt['index']]

    return datas